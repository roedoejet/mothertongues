from openpyxl import load_workbook
import pandas as pd
from mtd.parsers.utils import BaseParser
from mtd.exceptions import MissingResourceError, UnsupportedFiletypeError
from mtd.parsers.utils import ResourceManifest
from openpyxl.cell.cell import Cell
from typing import Dict, List, Tuple, Union
from string import ascii_letters
from tqdm import tqdm

class Parser(BaseParser):
    '''
    Parse data for MTD **TODO: test worksheet location. Skipheader in manifest skips first row. Location in manifest decides worksheet.

    :param ResourceManifest manifest: Manifest for parser
    :param str resource_path: path to file 
    '''
    def __init__(self, manifest: ResourceManifest, resource_path: str):
        self.manifest = manifest
        try:
            work_book = load_workbook(resource_path, data_only=True)
        except:
            raise UnsupportedFiletypeError(resource_path)
        if "location" in self.manifest:
            try:
                work_sheet = work_book[self.manifest["location"]]
            except KeyError:
                raise MissingResourceError(f"{self.manifest['location']} in {resource_path}")
        else:
            work_sheet = work_book.active
        if "skipheader" in self.manifest and self.manifest['skipheader']:
            min_row = 2
        else:
            min_row = 1
        self.resource = work_sheet.iter_rows(min_row=min_row)
        
        self.entry_template = self.manifest['targets']

    def col2num(self, col):
        ''' Turns letters into columns, A -> 1, B -> 2 etc...
            Some operating systems require this.
        '''
        num = 0
        for c in col:
            if c in ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def getCellValue(self, entry: Tuple[Cell, ...], col: str) -> str:
        ''' Given a tuple of OpenPyxl cells, return the value of the cell matching the column value for col
        '''
        for c in entry:
            if c.column == col or c.column == self.col2num(col):
                # Excel turns integers into floats, ie 1 -> 1.0 but we don't want that.
                if isinstance(c.value, float):
                    if c.value.is_integer():
                        return str(int(c.value))
                    else:
                        return str(c.value)
                else:
                    if c.value is not None:
                        return c.value
                    else:
                        return ''

    def resolve_targets(self) -> List[dict]:
        word_list = []
        for entry in tqdm(self.resource):
            word_list.append(self.fill_entry_template(self.entry_template, entry, self.getCellValue))
        return word_list

    def parse(self) -> Dict[str, Union[dict, pd.DataFrame]]:
        try:
            data = self.resolve_targets()
            return {"manifest": self.manifest, "data": pd.DataFrame(data)}
        except Exception as e:
            print(e)